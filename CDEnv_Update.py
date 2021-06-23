# -------------------------------------------------------------------------------
# Name:         CDEnv_Update
# Purpose:      Updates the Environmental Database with information from the
#               CD Projects Database, the Clients Database, and the Newspaper
#               Database.
#
# Author:      Justin Thornton
#
# Created:     06/17/2021
# Version:     0.1.1
# -------------------------------------------------------------------------------


from datetime import datetime
from openpyxl import workbook, load_workbook
print('...Update process intialized...')

# open and define workbooks
print('...Loading workbooks...')
project_wkb = load_workbook(
    "G:\\Shared drives\\Database\\Tables for Merge\\Working\\CD Projects.xlsx")
client_wkb = load_workbook(
    "G:\\Shared drives\\Database\\Tables for Merge\\Working\\Clients.xlsx")
np_wkb = load_workbook(
    "G:\\Shared drives\\Database\\Tables for Merge\\Working\\Newspapers_6.18.2020.xlsx")
env_wkb = load_workbook(
    "G:\\Shared drives\\Database\\Tables for Merge\\Working\\CD Env Review_5.12.2020_NEP DO NOT OVERWRITE.xlsx")

# define worksheets
print('...Loading worksheets...')
project_wks = project_wkb["CD Projects"]
client_wks = client_wkb["Clients"]
np_wks = np_wkb["Newspapers"]
env_wks = env_wkb["CD Env Review"]


def read_in_values(wks):
    """Stores worksheet values into a nested list

    Args:
        wks (worksheet): imported Worksheet
    """
    wks_array = [[value for value in row]for row in wks.values]
#   wks_array[row][column]
    return wks_array


def flatten_list(wks_array, item_index):
    """Creates a new one-dimensional list from a specific index (item_index) in a two-dimensional list
    (wks_array)

    Args:
        wks_array (list): two-dimensional list
        item_index (int): index of element in the nested list you want to flatten

    Returns:
        list: one-dimensional list
    """

    return [item[item_index] for item in wks_array]


def find_index(list1, search):
    """Identify the index of a specific element in a list

    Args:
        list1 (list): list to read
        search (any): search query

    Returns:
        int: the index of the element
    """

    return list1.index(search)


def check(env_row_check, wks_array_project, client_value):
    """If the contract row of the environmental database is missing
    a contract number, the function will search the CD Projects
    database for rows with identical clients. Each row is then compared
    with the environmental row (env_row_check) for matching elements.
    The row that has a match for every element will return the
    corresponding contract number.

    Args:
        env_row_check (list): the row in env database with missing contract number
        wks_array_project (list): the nested list to search (CD Projects database)
        client_value (str): the client

    Returns:
        str or int: the contract nummber
    """

    temp_list_project = [
        item for item in wks_array_project if item[0] == client_value]
    if bool(temp_list_project) is False:
        return
    for count, i in enumerate(temp_list_project):
        if env_row_check[15] == temp_list_project[count][6]:  # Grant Program
            if env_row_check[19] == temp_list_project[count][10]:  # Grant Amount
                if env_row_check[20] == temp_list_project[count][11]:  # Grant Match
                    if env_row_check[28] == temp_list_project[count][12]:  # Total Amount
                        if env_row_check[33] == temp_list_project[count][2]:  # Year
                            # contract number
                            return temp_list_project[count][1]


def from_projects_wks(starting_row, client_array, client_search, news_array, news_search):
    """Populates a new list with a length of 40. The function uses information
    from the CD Projects database.

    Args:
        starting_row (list): CD Projects row with pertinent data
        client_array (list): Nested list from Client database
        client_search (list): One dimensional list of clients
        news_array (list): Nested list from Newspaper database
        news_search (list): One dimensional list of newspapers

    Returns:
        list: New list of elements from CD Projects database, Client database
        and the Newspaper database.
    """
    new_row = []

    if starting_row[0] in client_search:
        row_index = find_index(client_search, starting_row[0])
        client_row = client_array[row_index]

        # Client, Contract
        new_row.extend(starting_row[0:2])

        from_client_wks(new_row, client_row)

        # Grant Program, Short Project Description, Long Project Description, Expr1018, Grant Amount, Match
        new_row.extend(starting_row[6:12])

        # blank spaces for posting info
        new_row.extend([" ", " ", " ", " ", " "])

        from_client_wks(new_row, client_row)

        # Main Contact Last
        new_row.append(starting_row[12])

        # Official Last
        new_row.append(starting_row[9])

        from_client_wks(new_row, client_row)

        # blank spaces for posting info
        new_row.extend([" ", " "])

        # Year
        new_row.append(starting_row[2])

        from_client_wks(new_row, client_row)
        from_np_wks(new_row, news_array, news_search)
        return new_row

    new_row.extend(starting_row[0:2])  # Client and Contract
    new_row.extend([" ", " ", " ", " ", " ", " ", " ",
                    " ", " ", " ", " ", " ", " ", " "])  # 12 blank spaces
    new_row.extend(starting_row[6:12])
    # 5 blank spaces for posting info
    new_row.extend([" ", " ", " ", " ", " "])
    new_row.extend([" ", " "])  # 2 blank spaces
    new_row.append(starting_row[12])
    new_row.append(starting_row[9])
    new_row.append('')
    new_row.extend([" ", " "])  # blank spaces for posting info
    new_row.append(starting_row[2])
    new_row.extend([" ", " ", " ", " "])
    from_np_wks(new_row, news_array, news_search)
    return new_row


def from_client_wks(new_row, client_row):
    """ The function uses information from the Client database
    to populate a new list.

    Args:
        new_row (list): New list for the CD Env database
        client_row (list): List of client information

    Returns:
        list: A list with information appended from the Client database.
    """

    if len(new_row) == 26:
        # City Hall/County Courthouse
        new_row.append(client_row[93])

        # County
        new_row.append(client_row[15])
        return new_row

    elif len(new_row) == 30:
        # Main Contact e-mail
        new_row.append(client_row[30])
        return new_row

    elif len(new_row) == 34:
        # Engineer Full Name, Short Eng
        new_row.extend(client_row[82:84])

        # Entity Type
        new_row.append(client_row[99])

        # Newspaper
        new_row.append(client_row[84])

        return new_row

    # Phys Address, City of, Zip Code, City
    new_row.extend(client_row[3:8])

    # Official Telephone
    new_row.append(client_row[24])

    # Official Last, Official First, Official Jr, Official Title
    new_row.extend(client_row[19:23])

    # Main Contact Last, Main Contact First, Main Contact Title
    new_row.extend(client_row[27:30])

    return new_row


def from_np_wks(new_row, np_array, np_line):
    """The function uses information fromthe Newspaper database
    to populate a new list.

    Args:
        new_row (list): New list for the CD Env database
        np_array (list): Nested list from Newspaper database
        np_line (list): List of Newspaper information

    Returns:
        list: A list with information appended from the Newspaper database.
    """

    if new_row[37] not in np_line:
        return new_row.extend([" ", " "])
    row_index = find_index(np_line, new_row[37])
    np_row = np_array[row_index]

    # Days Published, Deadline Date
    new_row.extend(np_row[3:5])
    return new_row


print('...defining intial arrays...')
env_array = read_in_values(env_wks)
project_array = read_in_values(project_wks)
client_array = read_in_values(client_wks)
news_array = read_in_values(np_wks)

print('...flattening arrays...')
env_contract_list = flatten_list(env_array, 1)
project_contract_list = flatten_list(project_array, 1)
client_by_client = flatten_list(client_array, 0)
news_by_news = flatten_list(news_array, 0)

# define trackers
total_contract_updates = 0
total_projects_added = 0

print('...Checking for existing projects with missing contract numbers...')
for index, env_rows in enumerate(env_wks.iter_rows(min_row=2, values_only=True)):
    if bool(env_wks.cell(row=index+2, column=1).value) is True and bool(env_wks.cell(row=index+2, column=2).value) is False:
        for row in env_wks.iter_rows(min_row=index+2, max_row=index+2, values_only=True):
            env_row = row
        env_wks.cell(row=index+2, column=2).value = check(env_row,
                                                          project_array, env_wks.cell(row=index+2, column=1).value)
        if env_wks.cell(row=index+2, column=2).value != None:
            total_contract_updates += 1
        env_contract_list.append(env_wks.cell(row=index+2, column=2).value)
# add a check for empty client info here

print('...Checking for missing projects...')
for i in project_contract_list:
    # add a check here for empty contracts in projects workbook
    if i not in env_contract_list:
        env_contract_list.append(i)
        row_index = find_index(project_contract_list, i)
        env_wks.append(
            from_projects_wks(project_array[row_index], client_array, client_by_client, news_array, news_by_news))
        total_projects_added += 1

print('...Saving Environmental Workbook...')
if total_contract_updates and total_projects_added > 0:
    env_wkb.save("CD Env Review_5.12.2020_NEP DO NOT OVERWRITE.xlsx")
    log_file = open("Env Changes.log", "a")
    date = datetime.utcnow().strftime('%m-%d-%y')
    log_file.write(f'\n\n{date}:\n')
    log_file.write(
        f'{total_contract_updates} Contract numbers were added to existing projects.\n{total_projects_added} Projects were added to the environmental database.\n\n')
    log_file.close()
print(f'{total_contract_updates} Contract numbers were added to existing projects.\n{total_projects_added} Projects were added to the environmental database.')
print('Update process complete.')
